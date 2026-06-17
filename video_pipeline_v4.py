# ============================================================
# ПАРАМЕТРЫ — меняешь только здесь
# ============================================================
CHUNK_DURATION  = 6          # длина чанка в секундах
CRF             = 23         # константа качества H.264 (стриминговый стандарт)
CODEC           = 'libx264'
MIN_SCALE       = 0.03       # нижняя граница ~3% = ~70p
TEST_MODE       = False

INPUT_DIR       = r'C:\video_exp'
OUTPUT_DIR      = r'C:\video_exp\results'
FFMPEG_PATH     = r'C:\ffmpeg\bin\ffmpeg.exe'

# ============================================================
# СЛОВАРЬ КОНТЕНТА
# ============================================================
CONTENT_LABELS = {
    'Netflix_Aerial':        'Аэросъёмка (дрон)',
    'Netflix_BarScene':      'Бар (интерьер, люди)',
    'Netflix_Dancers':       'Танцоры',
    'Netflix_DinnerScene':   'Ужин (интерьер, стол)',
    'Netflix_DrivingPOV':    'Вождение от первого лица',
    'Netflix_PierSeaside':   'Пирс у моря',
    'Netflix_RollerCoaster': 'Американские горки',
    'Netflix_WindAndNature': 'Ветер и природа',
}

# ============================================================
# 4 ПОСЛЕДОВАТЕЛЬНОСТИ
# ============================================================
SEQUENCES = {
    1: ['Netflix_Aerial',        'Netflix_BarScene',      'Netflix_Dancers',
        'Netflix_DinnerScene',   'Netflix_DrivingPOV',    'Netflix_PierSeaside',
        'Netflix_RollerCoaster', 'Netflix_WindAndNature'],
    2: ['Netflix_Dancers',       'Netflix_RollerCoaster', 'Netflix_WindAndNature',
        'Netflix_Aerial',        'Netflix_PierSeaside',   'Netflix_DinnerScene',
        'Netflix_BarScene',      'Netflix_DrivingPOV'],
    3: ['Netflix_RollerCoaster', 'Netflix_WindAndNature', 'Netflix_PierSeaside',
        'Netflix_BarScene',      'Netflix_Aerial',        'Netflix_Dancers',
        'Netflix_DrivingPOV',    'Netflix_DinnerScene'],
    4: ['Netflix_DinnerScene',   'Netflix_DrivingPOV',    'Netflix_BarScene',
        'Netflix_RollerCoaster', 'Netflix_WindAndNature', 'Netflix_Aerial',
        'Netflix_Dancers',       'Netflix_PierSeaside'],
}

# ============================================================
# ЗОНЫ ДЕГРАДАЦИИ
#
# Чанки  1-10: шаг 10%  длина 6 сек  100% → ~39%   4K→720p   00:00-01:00
#              незаметная зона — подтверждено просмотром на ноуте
# Чанки 11-28: шаг  6%  длина 5 сек   39% → ~12%   720p→280p 01:00-02:30
#              активная зона — 90 сек на 4 оценки (~22 сек на оценку)
#
# Обоснование:
#   - начало 10% не трогаем — по словам руководителя
#   - после 720p чанки короче (5 сек) — больше ступеней в активной зоне
#   - шаг 6% после 720p — переходы заметны но не скачут
#   - 90 сек активной зоны достаточно для 4 оценок с финалом ~15-20 сек
# ============================================================
def get_step(chunk_index):
    return 0.10   # единый шаг 10% для всех чанков

def calc_scale(chunk_index):
    scale = 1.0
    for i in range(chunk_index):
        scale *= (1 - get_step(i))
    return max(scale, MIN_SCALE)

# ============================================================
# 1. ИМПОРТЫ
# ============================================================
import ffmpeg
import os
import math
import pandas as pd
import subprocess
import sys
import shutil
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter

os.environ['PATH'] = os.path.dirname(FFMPEG_PATH) + os.pathsep + os.environ['PATH']

tmp_dir = os.path.join(OUTPUT_DIR, 'tmp')
for folder in [OUTPUT_DIR, tmp_dir]:
    if not os.path.exists(folder):
        os.makedirs(folder)
        print(f'  Создана папка: {folder}')
    else:
        print(f'  Папка уже есть: {folder}')

try:
    result = subprocess.run(['ffmpeg', '-version'], capture_output=True, text=True)
    ver = result.stdout.split('\n')[0]
    print(f'\n✓ FFmpeg найден: {ver}')
except Exception as e:
    print(f'✗ FFmpeg не найден: {e}')
    sys.exit(1)

# ============================================================
# 2. СКАНИРОВАНИЕ
# ============================================================
print(f'\nСканирую папку: {INPUT_DIR}')
source_files = []
for f in sorted(os.listdir(INPUT_DIR)):
    if f.endswith('.y4m'):
        full_path = os.path.join(INPUT_DIR, f)
        size_gb = os.path.getsize(full_path) / (1024**3)
        source_files.append(full_path)
        key   = next((k for k in CONTENT_LABELS if f.startswith(k)), f)
        label = CONTENT_LABELS.get(key, '—')
        print(f'  ✓ {f}  ({size_gb:.1f} ГБ)  [{label}]')

if not source_files:
    print(f'✗ y4m файлы не найдены в {INPUT_DIR}')
    sys.exit(1)
print(f'\nНайдено y4m файлов: {len(source_files)}')

# ============================================================
# 3. КОНВЕРТАЦИЯ y4m → mp4 (с проверкой)
# ============================================================
def convert_to_mp4(src, dst):
    (ffmpeg.input(src)
     .output(dst, vcodec=CODEC, crf=CRF,
             pix_fmt='yuv420p', an=None, preset='medium')
     .overwrite_output().run(quiet=True))

mp4_map = {}
print('\nКонвертация y4m → mp4:')
for src in source_files:
    basename = os.path.basename(src)
    key = next((k for k in CONTENT_LABELS if basename.startswith(k)), basename)
    dst = os.path.join(INPUT_DIR, basename.replace('.y4m', '_converted.mp4'))
    if os.path.exists(dst):
        size_mb = os.path.getsize(dst) / (1024**2)
        mp4_map[key] = dst
        print(f'  {basename} — уже есть ✓  ({size_mb:.0f} МБ)')
    else:
        print(f'  Конвертирую {basename} ...', end=' ', flush=True)
        convert_to_mp4(src, dst)
        size_mb = os.path.getsize(dst) / (1024**2)
        mp4_map[key] = dst
        print(f'✓  ({size_mb:.0f} МБ)')
print(f'\nИтого mp4: {len(mp4_map)} файлов')

# ============================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ============================================================
def get_video_info(path):
    probe  = ffmpeg.probe(path)
    stream = next(s for s in probe['streams'] if s['codec_type'] == 'video')
    return {
        'width':    int(stream['width']),
        'height':   int(stream['height']),
        'fps':      stream['r_frame_rate'],
        'duration': float(probe['format']['duration'])
    }

def get_bitrate_kbps(path):
    try:
        probe = ffmpeg.probe(path)
        return int(int(probe['format']['bit_rate']) / 1000)
    except Exception:
        return 0

# ============================================================
# ФОРМАТИРОВАНИЕ EXCEL
# ============================================================
# Цвета заголовков для каждой последовательности
SEQ_COLORS = {
    1: 'BDD7EE',   # голубой
    2: 'E2EFDA',   # зелёный
    3: 'FCE4D6',   # оранжевый
    4: 'EDE7F6',   # фиолетовый
}
HEADER_FONT  = Font(bold=True, size=10)
DATA_FONT    = Font(size=10)
CENTER       = Alignment(horizontal='center', vertical='center')
LEFT         = Alignment(horizontal='left',   vertical='center')

def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def write_excel(filepath, all_tables, sequences, content_labels,
                chunk_duration):
    """Создаёт Excel файл с отдельным листом для каждой последовательности
    и сводным листом."""
    wb = Workbook()
    wb.remove(wb.active)   # удаляем дефолтный пустой лист

    # колонки для вывода (без технических Ширина/Высота)
    cols = ['Чанк','Старт','Конец','Разрешение',
            'Масштаб_%','Битрейт_кбит','CRF','FPS','Контент']
    col_widths = [7, 8, 8, 14, 11, 14, 6, 8, 30]

    # ---- листы по последовательностям ----
    for seq_num, df in all_tables.items():
        ws = wb.create_sheet(title=f'Последовательность {seq_num}')
        color = SEQ_COLORS.get(seq_num, 'FFFFFF')
        fill  = PatternFill('solid', fgColor=color)

        # строка 1 — название последовательности
        ws.merge_cells(f'A1:{get_column_letter(len(cols))}1')
        c = ws.cell(1, 1,
            value=f'ПОСЛЕДОВАТЕЛЬНОСТЬ {seq_num}  |  '
                  f'Порядок: '
                  f'{" → ".join(content_labels.get(k,k) for k in sequences[seq_num])}')
        c.font      = Font(bold=True, size=11)
        c.fill      = fill
        c.alignment = LEFT
        c.border    = thin_border()

        # строка 2 — параметры кодирования
        ws.merge_cells(f'A2:{get_column_letter(len(cols))}2')
        c2 = ws.cell(2, 1,
            value=f'Кодек: H.264  |  CRF: 23  |  Длина чанка: {chunk_duration} сек  |  '
                  f'Шаг деградации: чанки 1-16 → 10%,  чанки 17+ → 5%')
        c2.font      = Font(italic=True, size=9)
        c2.fill      = PatternFill('solid', fgColor='F5F5F5')
        c2.alignment = LEFT
        c2.border    = thin_border()

        # строка 3 — заголовки столбцов
        headers_ru = ['Чанк','Старт','Конец','Разрешение',
                      'Масштаб, %','Битрейт, кбит/с','CRF','FPS','Контент']
        for col_idx, hdr in enumerate(headers_ru, 1):
            c = ws.cell(3, col_idx, value=hdr)
            c.font      = HEADER_FONT
            c.fill      = fill
            c.alignment = CENTER
            c.border    = thin_border()

        # данные
        alt_fill = PatternFill('solid', fgColor='FAFAFA')
        for row_idx, row in enumerate(df[cols].itertuples(index=False), 4):
            row_fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, val in enumerate(row, 1):
                c = ws.cell(row_idx, col_idx, value=val)
                c.font      = DATA_FONT
                c.alignment = CENTER if col_idx < len(cols) else LEFT
                c.border    = thin_border()
                if row_fill:
                    c.fill = row_fill

        # ширина столбцов
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 16
        ws.row_dimensions[3].height = 18
        ws.freeze_panes = 'A4'

    # ---- сводный лист ----
    ws_sum = wb.create_sheet(title='Сводная таблица', index=0)
    ws_sum.merge_cells(f'A1:{get_column_letter(len(cols)+1)}1')
    c = ws_sum.cell(1, 1, value='СВОДНАЯ ТАБЛИЦА — ВСЕ ПОСЛЕДОВАТЕЛЬНОСТИ')
    c.font = Font(bold=True, size=12)
    c.alignment = LEFT
    c.border = thin_border()

    sum_headers = ['Послед.'] + headers_ru
    for col_idx, hdr in enumerate(sum_headers, 1):
        c = ws_sum.cell(2, col_idx, value=hdr)
        c.font      = HEADER_FONT
        c.fill      = PatternFill('solid', fgColor='D9D9D9')
        c.alignment = CENTER
        c.border    = thin_border()

    row_idx = 3
    for seq_num, df in all_tables.items():
        fill = PatternFill('solid', fgColor=SEQ_COLORS.get(seq_num,'FFFFFF'))
        for row in df[cols].itertuples(index=False):
            ws_sum.cell(row_idx, 1, value=seq_num).border = thin_border()
            ws_sum.cell(row_idx, 1).fill      = fill
            ws_sum.cell(row_idx, 1).alignment = CENTER
            ws_sum.cell(row_idx, 1).font      = DATA_FONT
            for col_idx, val in enumerate(row, 2):
                c = ws_sum.cell(row_idx, col_idx, value=val)
                c.font      = DATA_FONT
                c.alignment = CENTER if col_idx < len(cols)+1 else LEFT
                c.border    = thin_border()
                c.fill      = fill
            row_idx += 1
        # пустая строка-разделитель между последовательностями
        row_idx += 1

    sum_widths = [12] + col_widths
    for col_idx, width in enumerate(sum_widths, 1):
        ws_sum.column_dimensions[get_column_letter(col_idx)].width = width
    ws_sum.freeze_panes = 'A3'

    wb.save(filepath)
    print(f'✓ Excel сохранён: {filepath}')

def save_csv_for_player(all_tables):
    """Сохраняет chunks_info.csv рядом с xlsx — для плеера."""
    csv_path = os.path.join(OUTPUT_DIR, 'chunks_info.csv')
    frames = []
    for seq_num, df in all_tables.items():
        df_copy = df.copy()
        df_copy.insert(0, 'Последовательность', seq_num)
        frames.append(df_copy)
    combined = pd.concat(frames, ignore_index=True)
    combined = combined.drop(columns=['Ширина','Высота'], errors='ignore')
    combined.to_csv(csv_path, index=False, encoding='utf-8-sig')
    print(f'✓ CSV для плеера сохранён: {csv_path}')

# ============================================================
# 4. ЦИКЛ ПО 4 ПОСЛЕДОВАТЕЛЬНОСТЯМ
# ============================================================
all_tables = {}

for seq_num, seq_keys in SEQUENCES.items():
    print(f'\n{"="*60}')
    print(f'  ПОСЛЕДОВАТЕЛЬНОСТЬ {seq_num}')
    print(f'  Порядок: {" → ".join(CONTENT_LABELS.get(k,k) for k in seq_keys)}')
    print(f'{"="*60}')

    # ----------------------------------------------------------
    # 4.1 СКЛЕЙКА reference_N.mp4 (с проверкой — если есть, пропускаем)
    # ----------------------------------------------------------
    reference_path = os.path.join(OUTPUT_DIR, f'reference_{seq_num}.mp4')

    if os.path.exists(reference_path):
        ref_size = os.path.getsize(reference_path) / (1024**2)
        print(f'\n  reference_{seq_num}.mp4 уже есть, пропускаем ✓  ({ref_size:.0f} МБ)')
    else:
        concat_list_path = os.path.join(tmp_dir, f'concat_{seq_num}.txt')
        with open(concat_list_path, 'w', encoding='utf-8') as f:
            for key in seq_keys:
                safe_path = mp4_map[key].replace('\\', '/')
                f.write(f"file '{safe_path}'\n")
        print(f'\n  Склейка reference_{seq_num}.mp4 ...', end=' ', flush=True)
        (ffmpeg.input(concat_list_path, format='concat', safe=0)
         .output(reference_path, vcodec=CODEC, crf=CRF,
                 pix_fmt='yuv420p', an=None)
         .overwrite_output().run(quiet=True))
        ref_size = os.path.getsize(reference_path) / (1024**2)
        print(f'✓  ({ref_size:.0f} МБ)')

    # ----------------------------------------------------------
    # 4.2 ПАРАМЕТРЫ ВИДЕО
    # ----------------------------------------------------------
    info      = get_video_info(reference_path)
    total_dur = info['duration']
    orig_w    = info['width']
    orig_h    = info['height']
    orig_fps  = info['fps']

    print(f'\n  Длительность : {total_dur:.1f} сек  ({total_dur/60:.1f} мин)')
    print(f'  Разрешение   : {orig_w}×{orig_h}')
    print(f'  FPS          : {orig_fps}')

    # ----------------------------------------------------------
    # 4.3 РАСЧЁТ ТАБЛИЦЫ ЧАНКОВ
    # ----------------------------------------------------------
    # переменная длина чанков: первые 10 по 6 сек, остальные по 5 сек
    # считаем сколько чанков влезает с учётом переменной длины
    n_chunks = 30   # фиксированно: 30×5 = 150 сек
    if TEST_MODE:
        n_chunks = 3
    if TEST_MODE:
        n_chunks = min(n_chunks, 3)
        print(f'\n  ⚠ ТЕСТОВЫЙ РЕЖИМ: только {n_chunks} чанка')

    chunks_per_file = int(20 / CHUNK_DURATION)

    # считаем стартовые позиции с учётом переменной длины чанков
    chunk_starts = []
    t = 0
    for i in range(n_chunks):
        chunk_starts.append(t)
        t += 5

    table = []
    for i in range(n_chunks):
        scale    = round(calc_scale(i), 6)
        w        = int(orig_w * scale) // 2 * 2
        h        = int(orig_h * scale) // 2 * 2
        dur_i    = 5
        start    = chunk_starts[i]
        # пропускаем чанк если старт выходит за конец файла
        if start >= total_dur:
            break
        # обрезаем длину последнего чанка если он выходит за конец
        dur_i    = min(dur_i, total_dur - start)
        end      = start + dur_i
        mm_s     = f'{int(start//60):02d}:{int(start%60):02d}'
        mm_e     = f'{int(end//60):02d}:{int(end%60):02d}'
        step_pct = int(get_step(i) * 100)
        dur_chunk = 5
        file_idx = min(i // chunks_per_file, len(seq_keys) - 1)
        content  = CONTENT_LABELS.get(seq_keys[file_idx], '—')

        table.append({
            'Чанк':         i + 1,
            'Старт':        mm_s,
            'Конец':        mm_e,
            'Длина_сек':    dur_chunk,
            'Старт_сек':    start,    # числовой старт для ffmpeg
            'Шаг_%':        step_pct,
            'Разрешение':   f'{w}×{h}',
            'Ширина':       w,
            'Высота':       h,
            'Масштаб_%':    round(scale * 100, 1),
            'CRF':          CRF,
            'Битрейт_кбит': 0,
            'FPS':          orig_fps,
            'Контент':      content,
        })

    df = pd.DataFrame(table)
    all_tables[seq_num] = df

    # ----------------------------------------------------------
    # 4.4 ОБРАБОТКА ЧАНКОВ
    # ----------------------------------------------------------
    processed_chunks = []
    print(f'\n  Обработка {n_chunks} чанков:')

    for row in table:
        i         = row['Чанк'] - 1
        start     = row['Старт_сек']   # берём из таблицы, не пересчитываем
        chunk_raw = os.path.join(tmp_dir, f's{seq_num}_chunk_{i+1:02d}_raw.mp4')
        chunk_out = os.path.join(tmp_dir, f's{seq_num}_chunk_{i+1:02d}_out.mp4')

        dur_i = row.get('Длина_сек', CHUNK_DURATION)
        # всегда пересоздаём chunk_raw — битый файл от упавшего запуска не мешает
        (ffmpeg.input(reference_path, ss=start, t=dur_i)
         .output(chunk_raw, vcodec=CODEC, crf=CRF,
                 pix_fmt='yuv420p', an=None)
         .overwrite_output().run(quiet=True))

        (ffmpeg.input(chunk_raw)
         .filter('scale', row['Ширина'], row['Высота'], **{'flags': 'lanczos'})
         .filter('scale', orig_w, orig_h, **{'flags': 'lanczos'})
         .output(chunk_out, vcodec=CODEC, crf=CRF,
                 pix_fmt='yuv420p', an=None, preset='medium')
         .overwrite_output().run(quiet=True))

        real_bitrate = get_bitrate_kbps(chunk_out)
        df.at[i, 'Битрейт_кбит'] = real_bitrate

        if os.path.exists(chunk_raw):
            os.remove(chunk_raw)

        processed_chunks.append(chunk_out)
        print(f'    ✓ Чанк {i+1:02d} | {row["Разрешение"]:>12} '
              f'| {row["Масштаб_%"]:>5}% | шаг {row["Шаг_%"]}% '
              f'| {real_bitrate} кбит/с')

    # ----------------------------------------------------------
    # 4.5 ФИНАЛЬНАЯ СКЛЕЙКА
    # ----------------------------------------------------------
    stimulus_path   = os.path.join(OUTPUT_DIR, f'test_stimulus_{seq_num}.mp4')
    final_list_path = os.path.join(tmp_dir, f'final_{seq_num}.txt')

    with open(final_list_path, 'w', encoding='utf-8') as f:
        for chunk in processed_chunks:
            safe = chunk.replace('\\', '/')
            f.write(f"file '{safe}'\n")

    print(f'\n  Финальная склейка test_stimulus_{seq_num}.mp4 ...', end=' ', flush=True)
    (ffmpeg.input(final_list_path, format='concat', safe=0)
     .output(stimulus_path, vcodec='copy')
     .overwrite_output().run(quiet=True))
    stim_size = os.path.getsize(stimulus_path) / (1024**2)
    print(f'✓  ({stim_size:.0f} МБ)')

    for chunk in processed_chunks:
        if os.path.exists(chunk):
            os.remove(chunk)

    # вывод таблицы в консоль (без столбца Зона)
    print(f'\n  === ТАБЛИЦА ЧАНКОВ — ПОСЛЕДОВАТЕЛЬНОСТЬ {seq_num} ===')
    print(df[['Чанк','Старт','Конец','Разрешение',
              'Масштаб_%','Битрейт_кбит','CRF','FPS','Контент']].to_string(index=False))

# ============================================================
# 5. СОХРАНЕНИЕ ДОКУМЕНТАЦИИ
# ============================================================

# -- sequences_description.txt --
desc_path = os.path.join(OUTPUT_DIR, 'sequences_description.txt')
with open(desc_path, 'w', encoding='utf-8') as f:
    f.write('=' * 60 + '\n')
    f.write('ТЕСТОВЫЕ ПОСЛЕДОВАТЕЛЬНОСТИ\n')
    f.write('=' * 60 + '\n\n')
    f.write(f'Параметры кодирования:\n')
    f.write(f'  Кодек          : H.264 (libx264)\n')
    f.write(f'  CRF            : {CRF}\n')
    f.write(f'  Длина чанка    : {CHUNK_DURATION} сек\n')
    f.write(f'  Чанков всего   : до 26\n')
    f.write(f'  Мин. масштаб   : {int(MIN_SCALE*100)}%\n\n')
    f.write('Схема деградации:\n')
    f.write('  30 чанков × 5 сек × шаг 10%  =  150 сек = 2:30\n')
    f.write('  100% → ~4%   (4096×2160 → ~164×86)\n\n')
    for seq_num, seq_keys in SEQUENCES.items():
        f.write('=' * 60 + '\n')
        f.write(f'test_stimulus_{seq_num}.mp4  /  reference_{seq_num}.mp4\n')
        f.write('Порядок фрагментов:\n')
        start_sec = 0
        for idx, key in enumerate(seq_keys):
            end_sec = start_sec + 20
            mm_s = f'{int(start_sec//60):02d}:{int(start_sec%60):02d}'
            mm_e = f'{int(end_sec//60):02d}:{int(end_sec%60):02d}'
            label = CONTENT_LABELS.get(key, key)
            f.write(f'  {idx+1}. {label:<30} [{mm_s} - {mm_e}]\n')
            start_sec = end_sec
        f.write('\n')
print(f'\n✓ sequences_description.txt сохранён')

# -- chunks_info.xlsx (Excel с форматированием) --
xlsx_path = os.path.join(OUTPUT_DIR, 'chunks_info.xlsx')
try:
    write_excel(xlsx_path, all_tables, SEQUENCES,
                CONTENT_LABELS, CHUNK_DURATION)
    save_csv_for_player(all_tables)
except ImportError:
    print('  ⚠ openpyxl не установлен — сохраняю как CSV')
    csv_path = os.path.join(OUTPUT_DIR, 'chunks_info.csv')
    frames = []
    for seq_num, df in all_tables.items():
        df_copy = df.copy()
        df_copy.insert(0, 'Последовательность', seq_num)
        frames.append(df_copy)
    combined = pd.concat(frames, ignore_index=True)
    combined = combined.drop(columns=['Ширина','Высота'])
    combined.to_csv(csv_path, index=False, encoding='utf-8-sig')
    print(f'✓ CSV сохранён: {csv_path}')

# ============================================================
# 6. ОЧИСТКА tmp
# ============================================================
shutil.rmtree(tmp_dir)
print(f'✓ Папка tmp очищена')

# ============================================================
# 7. ИТОГОВЫЙ ОТЧЁТ
# ============================================================
print('\n' + '=' * 60)
print('                    ГОТОВО')
print('=' * 60)
for seq_num in SEQUENCES:
    ref_p   = os.path.join(OUTPUT_DIR, f'reference_{seq_num}.mp4')
    stim_p  = os.path.join(OUTPUT_DIR, f'test_stimulus_{seq_num}.mp4')
    ref_mb  = os.path.getsize(ref_p)  / (1024**2) if os.path.exists(ref_p)  else 0
    stim_mb = os.path.getsize(stim_p) / (1024**2) if os.path.exists(stim_p) else 0
    print(f'  reference_{seq_num}.mp4      {ref_mb:.0f} МБ  (без изменений)')
    print(f'  test_stimulus_{seq_num}.mp4  {stim_mb:.0f} МБ  (новая схема деградации)')
print(f'  sequences_description.txt')
print(f'  chunks_info.xlsx')
print('=' * 60)
print(f'\n  Кодек          : H.264  CRF={CRF}')
print(f'  Длина чанка    : {CHUNK_DURATION} сек')
print(f'  Чанков         : до 53')
print(f'  Схема шагов    : 10% (чанки 1-16) → 5% (чанки 17+)')
print(f'  Диапазон       : 100% → {int(MIN_SCALE*100)}%')
print(f'\n  Файлы сохранены: {OUTPUT_DIR}')
