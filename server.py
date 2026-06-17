# =============================================================================
# Листинг А.2 — Серверный модуль системы субъективной оценки качества видео
#
# Назначение: реализует локальный веб-сервер на основе Flask для проведения
# субъективного эксперимента по методике ITU-R BT.500. Обеспечивает
# управление сеансом воспроизведения видеопоследовательностей и сбор
# субъективных оценок качества от группы участников через мобильные
# устройства, подключённые по протоколу Wi-Fi.
#
# Структура модуля:
#   1. Настраиваемые параметры
#   2. Разделяемое состояние (Flask-поток ↔ консоль организатора)
#   3. Вспомогательные функции
#   4. Работа с данными (загрузка, поиск, генерация документации)
#   5. Управление разрешением дисплея (Windows API)
#   6. HTML-страницы веб-интерфейса
#   7. Маршруты Flask HTTP API
#   8. Воспроизведение видео (FFplay)
#   9. Консольный интерфейс организатора
#  10. Запись результатов (CSV, XLSX)
#  11. Главная функция и точка входа
#
# Нормативные документы:
#   ITU-R BT.500-14 — методология субъективной оценки качества видео
#   ГОСТ 19.401-78 ЕСПД — требования к оформлению текстов программ
# =============================================================================

# ============================================================
# 1. НАСТРАИВАЕМЫЕ ПАРАМЕТРЫ
# Все пути и константы, изменяемые при переносе на другой ПК,
# сосредоточены в данном блоке.
# ============================================================
VIDEO_DIR   = r'C:\video_exp\results'       # каталог с тестовыми видеофайлами
FFPLAY_PATH = r'C:\ffmpeg\bin\ffplay.exe'   # путь к исполняемому файлу FFplay
CHUNKS_CSV  = r'C:\video_exp\results\chunks_info.csv'  # таблица метаданных чанков
RESULTS_DIR = r'C:\video_exp\results'       # корневой каталог результатов сессий
RESULTS_CSV = ''   # путь к CSV текущей сессии; задаётся динамически в main()
PORT        = 5000 # номер TCP-порта веб-сервера

# Соответствие порядковых номеров видео именам файлов тестовых стимулов
VIDEO_MAP  = {1:'test_stimulus_1.mp4', 2:'test_stimulus_2.mp4',
              3:'test_stimulus_3.mp4', 4:'test_stimulus_4.mp4'}

# Шкала субъективных оценок качества по ITU-R BT.500 (от «отлично» до «плохо»)
MOS_SCORES = [5, 4, 3, 2]

# Последовательность форматов экрана в порядке предъявления
FORMATS    = ['4K', '1080', '720']

# Соответствие обозначений форматов физическим разрешениям (ширина × высота, пикс.)
RESOLUTION_MAP = {
    '4K':   (3840, 2160),
    '1080': (1920, 1080),
    '720':  (1280, 720),
}

# ============================================================
# 2. ИМПОРТ БИБЛИОТЕК
# ============================================================
import os, sys, time, csv, socket, threading, subprocess, random
import pandas as pd
from datetime import datetime
from flask import Flask, request, jsonify, render_template_string

try:
    import qrcode, io, base64  # генерация QR-кода ссылки для подключения
    HAS_QR = True
except ImportError:
    HAS_QR = False  # при отсутствии библиотеки QR заменяется текстовой ссылкой

# ============================================================
# 3. РАЗДЕЛЯЕМОЕ СОСТОЯНИЕ
# Словарь S содержит переменные, к которым обращаются одновременно
# Flask-поток (обработка HTTP-запросов) и консоль организатора.
# Доступ защищён реентрантным мьютексом RLock, позволяющим одному
# потоку повторно захватывать блокировку без взаимной блокировки.
# ============================================================
S = {
    'active':        False,  # признак активного воспроизведения (кнопки разблокированы)
    'exp_num':       0,      # порядковый номер текущего эксперимента в сессии
    'total_exp':     0,      # общее число экспериментов в сессии
    'video_num':     None,   # номер воспроизводимого видео (1–4)
    'screen_format': None,   # текущий формат экрана ('4K' / '1080' / '720')
    'timer_start':   None,   # метка времени начала воспроизведения (time.time())
    'exp_version':   0,      # счётчик версий; инкрементируется при каждом новом/повторном эксперименте
    'users':         {},     # данные участников: {uid: {ip, online, taps, joined, last_seen}}
    'ip_to_uid':     {},     # обратный индекс IP-адрес → uid
    'events':        [],     # очередь журнальных событий для вывода в консоль
    'chunks_df':     None,   # DataFrame с метаданными чанков (из chunks_info.csv)
    'lock': threading.RLock(),  # реентрантный мьютекс для потокобезопасного доступа
}

# ============================================================
# 4. ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ============================================================
def ts():
    """Возвращает текущее время в виде строки ЧЧ:ММ:СС."""
    return datetime.now().strftime('%H:%M:%S')

def mmss(sec):
    """Преобразует длительность в секундах в строку формата ММ:СС."""
    return f'{int(sec//60):02d}:{int(sec%60):02d}'

def log(msg):
    """Добавляет событие в очередь журнала; очередь ограничена 50 записями."""
    with S['lock']:
        S['events'].append(f'  [{ts()}] {msg}')
        if len(S['events']) > 50:
            S['events'].pop(0)

def exp_key():
    """Формирует уникальный ключ текущего эксперимента для адресации массива оценок.
    Формат: v<номер_видео>_<формат>_e<версия>."""
    return f"v{S['video_num']}_{S['screen_format']}_e{S['exp_version']}"

def ask(prompt, options):
    """Запрашивает ввод организатора; повторяет запрос до получения допустимого значения."""
    while True:
        val = input(prompt).strip().lower()
        if val in options:
            return val
        print(f'  Введите одно из: {", ".join(options)}')

# ============================================================
# 5. РАБОТА С ДАННЫМИ
# ============================================================
def load_chunks():
    """Загружает таблицу метаданных чанков из chunks_info.csv.
    Таблица используется для сопоставления таймкода нажатия кнопки
    с параметрами чанка (разрешение, масштаб, битрейт, FPS)."""
    try:
        S['chunks_df'] = pd.read_csv(CHUNKS_CSV, encoding='utf-8-sig')
        print('  ✓ chunks_info.csv загружен')
        generate_sequences_txt()  # однократная генерация описания последовательностей
    except Exception as e:
        print(f'  ⚠ chunks_info.csv не загружен: {e}')

def generate_sequences_txt():
    """Формирует файл sequences_description.txt с описанием порядка фрагментов
    в каждой тестовой последовательности. При повторном запуске не перезаписывается."""
    txt_path = os.path.join(RESULTS_DIR, 'sequences_description.txt')
    if os.path.exists(txt_path):
        print('  ✓ sequences_description.txt уже есть')
        return
    df = S['chunks_df']
    if df is None:
        return
    try:
        sequences = df['Последовательность'].unique()
        with open(txt_path, 'w', encoding='utf-8') as f:
            f.write('=' * 60 + '\n')
            f.write('ТЕСТОВЫЕ ПОСЛЕДОВАТЕЛЬНОСТИ\n')
            f.write('=' * 60 + '\n\n')
            row0 = df.iloc[0]
            f.write(f'Параметры кодирования:\n'
                    f'  Кодек          : H.264 (libx264)\n'
                    f'  CRF            : {row0.get("CRF", 23)}\n'
                    f'  Длина чанка    : 5 сек (чанки 1-9: 6 сек)\n'
                    f'  Чанков всего   : 30\n'
                    f'  FPS            : {row0.get("FPS", "60/1")}\n\n'
                    f'Схема деградации:\n'
                    f'  30 чанков × 5 сек × шаг 10%  =  150 сек = 2:30\n'
                    f'  100% → ~4%   (4096×2160 → ~164×86)\n\n')
            for seq_num in sorted(sequences):
                seq_df = df[df['Последовательность'] == seq_num].sort_values('Чанк')
                last_end = seq_df.iloc[-1]['Конец']
                f.write('=' * 60 + '\n')
                f.write(f'test_stimulus_{seq_num}.mp4  /  reference_{seq_num}.mp4\n')
                f.write(f'Длительность   : {last_end}\n')
                f.write('Порядок фрагментов (по контенту):\n')
                seen = {}
                for _, row in seq_df.iterrows():
                    cont = row.get('Контент', '—')
                    if cont not in seen:
                        seen[cont] = {'start': row['Старт'], 'end': row['Конец']}
                    else:
                        seen[cont]['end'] = row['Конец']
                for clip_idx, (cont, times) in enumerate(seen.items(), 1):
                    f.write(f'  {clip_idx}. {cont:<32} [{times["start"]} - {times["end"]}]\n')
                f.write('\n')
        print('  ✓ sequences_description.txt создан')
    except Exception as e:
        print(f'  ⚠ sequences_description.txt не создан: {e}')

def find_chunk(seq_num, tc_sec):
    """Находит строку метаданных чанка по номеру последовательности и таймкоду.
    Возвращает словарь с полями Разрешение, Масштаб_%, Битрейт_кбит, FPS."""
    df = S['chunks_df']
    if df is None:
        return {}
    try:
        for _, row in df[df['Последовательность'] == seq_num].iterrows():
            ss = str(row['Старт']).split(':')
            es = str(row['Конец']).split(':')
            if int(ss[0])*60+int(ss[1]) <= tc_sec < int(es[0])*60+int(es[1]):
                return row.to_dict()
    except:
        pass
    return {}

def max_uid_in_csv():
    """Определяет максимальный номер участника в накопленных результатах.
    Используется для информирования организатора о диапазоне номеров новой сессии."""
    if not os.path.exists(RESULTS_CSV):
        return 0
    try:
        df = pd.read_csv(RESULTS_CSV, encoding='utf-8-sig')
        if 'Испытуемый' in df.columns and len(df):
            return int(df['Испытуемый'].max())
    except:
        pass
    return 0

def get_local_ip():
    """Определяет IP-адрес узла в локальной сети (Wi-Fi или LAN).
    Фильтрует виртуальные адаптеры Windows (Hyper-V, мобильный хотспот)."""
    candidates = []
    try:
        for info in socket.getaddrinfo(socket.gethostname(), None):
            ip = info[4][0]
            if ':' in ip or ip.startswith(('127.','10.254.','169.254.')):
                continue  # пропускаем IPv6, loopback и виртуальные адаптеры
            candidates.append(ip)
    except:
        pass
    for ip in candidates:
        if ip.startswith('192.168.'):  # предпочтение домашней/офисной сети
            return ip
    for ip in candidates:
        if ip.startswith('10.'):       # затем корпоративная сеть
            return ip
    if candidates:
        return candidates[0]
    try:                               # резервный способ через UDP-сокет
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        return s.getsockname()[0]
    except:
        return '127.0.0.1'

# ============================================================
# 6. УПРАВЛЕНИЕ РАЗРЕШЕНИЕМ ДИСПЛЕЯ (Windows API)
# Используется функция ChangeDisplaySettingsW из библиотеки user32.dll.
# Перед применением выполняется тестовый вызов с флагом CDS_TEST.
# ============================================================
# Параметры второго монитора (AP7_Titanium)
# Определены командой: [System.Windows.Forms.Screen]::AllScreens | Format-List *
#   DeviceName: \\.\DISPLAY2, Bounds: X=1280, Y=12, Width=3840, Height=2160
SECOND_MONITOR_DEVICE = r'\\.\DISPLAY2'  # системное имя второго монитора
SECOND_MONITOR_X      = 1280               # X-координата левого края (= ширина ноута)
SECOND_MONITOR_Y      = 12                 # Y-координата верхнего края

def get_second_monitor_device():
    """Возвращает системное имя второго монитора (DISPLAY2).
    Имя определено заранее через PowerShell и задано константой."""
    return SECOND_MONITOR_DEVICE

def check_second_monitor():
    """Проверяет подключение второго монитора через EnumDisplaySettingsW.
    Возвращает True если DISPLAY2 активен, False если не найден."""
    try:
        import ctypes
        from ctypes import wintypes

        class DEVMODE(ctypes.Structure):
            _fields_ = [
                ('dmDeviceName',       ctypes.c_wchar*32),
                ('dmSpecVersion',      wintypes.WORD),
                ('dmDriverVersion',    wintypes.WORD),
                ('dmSize',             wintypes.WORD),
                ('dmDriverExtra',      wintypes.WORD),
                ('dmFields',           wintypes.DWORD),
                ('dmPositionX',        wintypes.LONG),
                ('dmPositionY',        wintypes.LONG),
                ('dmDisplayOrientation',  wintypes.DWORD),
                ('dmDisplayFixedOutput',  wintypes.DWORD),
                ('dmColor',            wintypes.SHORT),
                ('dmDuplex',           wintypes.SHORT),
                ('dmYResolution',      wintypes.SHORT),
                ('dmTTOption',         wintypes.SHORT),
                ('dmCollate',          wintypes.SHORT),
                ('dmFormName',         ctypes.c_wchar*32),
                ('dmLogPixels',        wintypes.WORD),
                ('dmBitsPerPel',       wintypes.DWORD),
                ('dmPelsWidth',        wintypes.DWORD),
                ('dmPelsHeight',       wintypes.DWORD),
                ('dmDisplayFlags',     wintypes.DWORD),
                ('dmDisplayFrequency', wintypes.DWORD),
                ('dmICMMethod',        wintypes.DWORD),
                ('dmICMIntent',        wintypes.DWORD),
                ('dmMediaType',        wintypes.DWORD),
                ('dmDitherType',       wintypes.DWORD),
                ('dmReserved1',        wintypes.DWORD),
                ('dmReserved2',        wintypes.DWORD),
                ('dmPanningWidth',     wintypes.DWORD),
                ('dmPanningHeight',    wintypes.DWORD),
            ]
        dm = DEVMODE()
        dm.dmSize = ctypes.sizeof(DEVMODE)
        # если EnumDisplaySettingsW вернул данные — монитор подключён
        result = ctypes.windll.user32.EnumDisplaySettingsW(
            SECOND_MONITOR_DEVICE, 0, ctypes.byref(dm))
        return result != 0
    except:
        return False

def change_resolution(fmt):
    """Изменяет разрешение второго монитора (AP7_Titanium) через Windows API.
    Использует ChangeDisplaySettingsExW для адресации конкретного дисплея.
    Возвращает кортеж (успех: bool, сообщение: str)."""
    if fmt not in RESOLUTION_MAP:
        return False, f'Неизвестный формат {fmt}'
    w, h = RESOLUTION_MAP[fmt]
    try:
        import ctypes
        from ctypes import wintypes

        class DEVMODE(ctypes.Structure):
            _fields_ = [
                ('dmDeviceName',       ctypes.c_wchar*32),
                ('dmSpecVersion',      wintypes.WORD),
                ('dmDriverVersion',    wintypes.WORD),
                ('dmSize',             wintypes.WORD),
                ('dmDriverExtra',      wintypes.WORD),
                ('dmFields',           wintypes.DWORD),
                ('dmPositionX',        wintypes.LONG),
                ('dmPositionY',        wintypes.LONG),
                ('dmDisplayOrientation',  wintypes.DWORD),
                ('dmDisplayFixedOutput',  wintypes.DWORD),
                ('dmColor',            wintypes.SHORT),
                ('dmDuplex',           wintypes.SHORT),
                ('dmYResolution',      wintypes.SHORT),
                ('dmTTOption',         wintypes.SHORT),
                ('dmCollate',          wintypes.SHORT),
                ('dmFormName',         ctypes.c_wchar*32),
                ('dmLogPixels',        wintypes.WORD),
                ('dmBitsPerPel',       wintypes.DWORD),
                ('dmPelsWidth',        wintypes.DWORD),
                ('dmPelsHeight',       wintypes.DWORD),
                ('dmDisplayFlags',     wintypes.DWORD),
                ('dmDisplayFrequency', wintypes.DWORD),
                ('dmICMMethod',        wintypes.DWORD),
                ('dmICMIntent',        wintypes.DWORD),
                ('dmMediaType',        wintypes.DWORD),
                ('dmDitherType',       wintypes.DWORD),
                ('dmReserved1',        wintypes.DWORD),
                ('dmReserved2',        wintypes.DWORD),
                ('dmPanningWidth',     wintypes.DWORD),
                ('dmPanningHeight',    wintypes.DWORD),
            ]

        # используем заранее определённое имя второго монитора
        device = get_second_monitor_device()

        dm = DEVMODE()
        dm.dmSize             = ctypes.sizeof(DEVMODE)
        dm.dmPelsWidth        = w
        dm.dmPelsHeight       = h
        dm.dmDisplayFrequency = 60
        dm.dmFields           = 0x80000 | 0x100000 | 0x400000

        if device:
            # меняем конкретный монитор через Ex-версию функции
            res_test  = ctypes.windll.user32.ChangeDisplaySettingsExW(
                device, ctypes.byref(dm), None, 0x02, None)  # CDS_TEST
            if res_test != 0:
                return False, f'{w}×{h} не поддерживается монитором DISPLAY2'
            res_apply = ctypes.windll.user32.ChangeDisplaySettingsExW(
                device, ctypes.byref(dm), None, 0x01, None)  # CDS_UPDATEREGISTRY
            if res_apply == 0:
                return True, f'Установлено {w}×{h} @ 60 Гц на DISPLAY2'
            return False, f'Ошибка применения разрешения (код {res_apply})'
        else:
            # fallback на основной дисплей
            if ctypes.windll.user32.ChangeDisplaySettingsW(ctypes.byref(dm), 0x02) != 0:
                return False, f'{w}×{h} не поддерживается дисплеем'
            if ctypes.windll.user32.ChangeDisplaySettingsW(ctypes.byref(dm), 0x01) == 0:
                return True, f'Установлено {w}×{h} @ 60 Гц'
            return False, 'Ошибка применения разрешения'
    except Exception as ex:
        return False, f'Ошибка: {ex}'

# ============================================================
# 7. HTML-СТРАНИЦЫ ВЕБ-ИНТЕРФЕЙСА
# PAGE_USER — страница участника: форма ввода номера и кнопка оценки.
# PAGE_QR   — страница с QR-кодом для отображения на большом экране.
# ============================================================
PAGE_USER = r"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1,user-scalable=no">
<title>Оценка качества видео</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:Arial,sans-serif;background:#0f0f1a;color:#fff;
     display:flex;flex-direction:column;align-items:center;
     justify-content:center;height:100vh;overflow:hidden;user-select:none}
#login{text-align:center;padding:30px}
#login h2{font-size:1.8em;color:#a0c4ff;margin-bottom:10px}
#login p{color:#888;margin-bottom:25px}
#login input{font-size:2.5em;width:140px;text-align:center;padding:12px;
             border-radius:12px;border:2px solid #a0c4ff;
             background:#16213e;color:#fff;outline:none}
#login button{display:block;margin:20px auto 0;font-size:1.3em;
              padding:14px 50px;background:#a0c4ff;color:#0f0f1a;
              border:none;border-radius:12px;cursor:pointer;font-weight:bold}
#uid_err{color:#e74c3c;font-size:0.9em;margin-top:10px;min-height:20px}
#main{display:none;width:100%;height:100%;flex-direction:column;
      align-items:center;justify-content:center;padding:20px}
#main.visible{display:flex !important}
#exp_info{font-size:1em;color:#a0c4ff;margin-bottom:15px;text-align:center;
          background:#16213e;padding:10px 20px;border-radius:10px;width:90vw}
#btn{width:90vw;height:50vh;border-radius:25px;border:none;
     font-size:1.6em;font-weight:bold;cursor:pointer;transition:all 0.15s;
     -webkit-tap-highlight-color:transparent}
#btn.waiting{background:#2a2a3e;color:#555;cursor:not-allowed}
#btn.ready  {background:#27ae60;color:#fff;box-shadow:0 0 40px #27ae6088}
#btn.confirm{background:#e67e22;color:#fff}
#btn.done   {background:#1a3a2a;color:#2ecc71;cursor:not-allowed}
#btn.offline{background:#2a2a3e;color:#555;cursor:not-allowed}
#dots{display:flex;gap:12px;margin-top:20px}
.dot{width:22px;height:22px;border-radius:50%;background:#2a2a3e;transition:background 0.3s}
.dot.on{background:#27ae60}
#status_msg{margin-top:15px;font-size:0.95em;color:#888;text-align:center;min-height:24px}
#uid_badge{position:fixed;top:12px;right:16px;background:#16213e;
           color:#a0c4ff;padding:6px 14px;border-radius:20px;font-size:0.85em}
</style>
</head>
<body>
<div id="login">
  <h2>Оценка качества видео</h2>
  <p>Введите ваш номер участника</p>
  <input type="number" id="uid_inp" min="0" max="999" placeholder="№" autofocus>
  <div id="uid_err"></div>
  <button onclick="doLogin()">Войти</button>
</div>
<div id="main">
  <div id="uid_badge">Участник №<span id="uid_show"></span></div>
  <div id="exp_info">Ожидаем начала...</div>
  <button id="btn" class="waiting" onclick="handleTap()">Ожидайте начала эксперимента</button>
  <div id="dots">
    <div class="dot" id="d0"></div><div class="dot" id="d1"></div>
    <div class="dot" id="d2"></div><div class="dot" id="d3"></div>
  </div>
  <div id="status_msg"></div>
</div>
<script>
let uid=null, tapCount=0, pending=false;
const MOS_SCORES=[5,4,3,2];
let lastExpVer=-1, joinedActive=false, lastState=null;

// Обработчик клавиши Enter в поле ввода номера участника
document.getElementById('uid_inp').addEventListener('keydown',e=>{if(e.key==='Enter')doLogin()});

// Переключение с экрана входа на экран эксперимента
function showMain(){
  document.getElementById('login').style.cssText='display:none!important';
  const m=document.getElementById('main');
  m.style.cssText='display:flex!important;width:100%;height:100%;flex-direction:column;align-items:center;justify-content:center;padding:20px';
  m.classList.add('visible');
}

// Регистрация участника на сервере (GET /register?uid=N)
function doLogin(){
  const v=parseInt(document.getElementById('uid_inp').value);
  const err=document.getElementById('uid_err');
  if(isNaN(v)||v<0||v>999){err.textContent='Введите число от 0 до 999';return;}
  err.textContent='Подключение...'; err.style.color='#a0c4ff';
  const xhr=new XMLHttpRequest();
  xhr.open('GET','/register?uid='+v); xhr.timeout=10000;
  xhr.onload=function(){
    try{
      const d=JSON.parse(xhr.responseText);
      if(d.ok){
        uid=v;
        document.getElementById('uid_show').textContent=uid;
        tapCount=parseInt(d.tap_count)||0;
        updateDots(); showMain(); poll();
      } else {err.style.color='#e74c3c';err.textContent=d.error||'Ошибка';}
    }catch(e){err.style.color='#e74c3c';err.textContent='Ошибка ответа';}
  };
  xhr.onerror  =function(){err.style.color='#e74c3c';err.textContent='Нет связи с сервером';};
  xhr.ontimeout=function(){err.style.color='#e74c3c';err.textContent='Таймаут — проверьте Wi-Fi';};
  xhr.send();
}

// Обработка нажатия кнопки оценки; флаг pending предотвращает повторные нажатия
function handleTap(){
  if(!joinedActive||tapCount>=4||pending) return;
  pending=true;
  const btn=document.getElementById('btn');
  btn.className='confirm'; btn.textContent='Фиксируем...';
  sendTap();
}

// Отправка оценки на сервер (POST /tap) и обновление состояния кнопки
function sendTap(){
  const xhr=new XMLHttpRequest();
  xhr.open('POST','/tap');
  xhr.setRequestHeader('Content-Type','application/json');
  xhr.timeout=8000;
  xhr.onload=function(){
    try{
      const d=JSON.parse(xhr.responseText);
      if(d.ok){tapCount=d.tap_count; updateDots(); setStatus('Оценка '+d.mos+' — '+d.timecode);}
    }catch(e){}
    pending=false; updateBtn();  // снимаем блокировку и обновляем кнопку
  };
  xhr.onerror  =function(){pending=false; updateBtn();};
  xhr.ontimeout=function(){pending=false; updateBtn();};
  xhr.send(JSON.stringify({uid:uid}));
}

// Централизованное обновление визуального состояния кнопки
// на основе флагов active, joinedActive, tapCount и номера эксперимента
function updateBtn(){
  if(pending) return;  // не изменяем кнопку во время ожидания ответа сервера
  const btn=document.getElementById('btn');
  const act=lastState&&lastState.active;
  if(tapCount>=4){
    // все 4 оценки зафиксированы
    const isLast=lastState&&lastState.exp_num>=lastState.total_exp&&!act;
    btn.className='done';
    btn.textContent=isLast
      ?'Сессия завершена. Спасибо за участие!'
      :'✓ Все оценки зафиксированы';
    if(isLast) setStatus('Эту страницу можно закрыть'); else setStatus('Ждите следующего эксперимента');
  } else if(act&&joinedActive){
    // видео воспроизводится, участник подключился до старта
    btn.className='ready';
    btn.textContent=tapCount===0
      ?'Нажмите чтобы зафиксировать эталонное качество (оценка 5)'
      :'Нажмите когда заметите ухудшение качества (оценка '+MOS_SCORES[tapCount]+')';
  } else if(act&&!joinedActive){
    // видео воспроизводится, но участник подключился после старта
    btn.className='offline'; btn.textContent='Эксперимент уже идёт — ждите следующего';
    setStatus('Ваши оценки за этот эксперимент = None');
  } else {
    // видео не воспроизводится — режим ожидания
    const num=lastState?lastState.exp_num:0;
    const tot=lastState?lastState.total_exp:0;
    btn.className='waiting';
    if(num>0&&num>=tot)
      btn.textContent='Сессия завершена. Спасибо за участие!';
    else
      btn.textContent=num>0?`Ожидайте эксперимент ${num}...`:'Ожидайте начала...';
    joinedActive=false;
  }
}

// Периодический опрос состояния сервера каждые 800 мс (polling)
function poll(){
  fetch('/state?uid='+uid).then(r=>r.json()).then(d=>{
    // при смене версии эксперимента сбрасываем локальный прогресс участника
    if(d.exp_version!==lastExpVer){
      lastExpVer=d.exp_version; tapCount=d.tap_count||0;
      joinedActive=false; pending=false; updateDots(); setStatus('');
    }
    // отображаем информацию о текущем эксперименте
    const info=document.getElementById('exp_info');
    info.textContent=d.exp_num>0
      ?`Эксперимент ${d.exp_num}/${d.total_exp} │ Видео ${d.video_num} │ ${d.screen_format}`
      :'Ожидаем начала...';
    // фиксируем факт подключения до старта воспроизведения
    if(d.active&&!joinedActive&&tapCount===0) joinedActive=true;
    lastState=d;
    updateBtn();
    setTimeout(poll,800);
  }).catch(()=>setTimeout(poll,2000));
}

// Возобновление опроса при возврате на вкладку после фоновой паузы браузера
document.addEventListener('visibilitychange',function(){
  if(document.visibilityState==='visible'&&uid!==null) poll();
});

function updateDots(){
  for(let i=0;i<4;i++)
    document.getElementById('d'+i).className='dot'+(i<tapCount?' on':'');
}
function setStatus(m){document.getElementById('status_msg').textContent=m;}
</script>
</body>
</html>"""

PAGE_QR = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<title>QR — подключение</title>
<style>
body{background:#0f0f1a;display:flex;flex-direction:column;align-items:center;
     justify-content:center;height:100vh;color:#fff;font-family:Arial,sans-serif;text-align:center}
h1{font-size:2.5em;color:#a0c4ff;margin-bottom:8px}
p{font-size:1.2em;color:#888;margin-bottom:30px}
img{border-radius:20px;padding:18px;background:#fff;box-shadow:0 0 60px #a0c4ff44}
.url{margin-top:28px;font-size:1.8em;color:#27ae60;
     background:#16213e;padding:16px 36px;border-radius:12px}
</style>
</head>
<body>
<h1>Сканируйте QR-код</h1>
<p>или введите адрес в браузере телефона</p>
{% if qr_img %}<img src="data:image/png;base64,{{ qr_img }}" width="380" height="380">
{% else %}<p style="color:#e67e22">⚠ pip install qrcode[pil]</p>{% endif %}
<div class="url">{{ url }}</div>
</body>
</html>"""

# ============================================================
# 8. МАРШРУТЫ FLASK HTTP API
# ============================================================
app = Flask(__name__)

# Заголовки CORS разрешают запросы с любых устройств локальной сети
@app.after_request
def add_cors(response):
    response.headers['Access-Control-Allow-Origin']  = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response

@app.route('/ping')
def ping():
    """Диагностический маршрут для проверки доступности сервера с клиента."""
    return app.response_class('{"ok":true,"msg":"server ok"}', mimetype='application/json')

@app.route('/')
def index():
    """Возвращает страницу участника с формой входа и кнопкой оценки."""
    return render_template_string(PAGE_USER)

@app.route('/qr')
def qr_page():
    """Возвращает страницу с QR-кодом ссылки на сервер для большого экрана."""
    url    = f'http://{get_local_ip()}:{PORT}'
    qr_img = None
    if HAS_QR:
        buf = io.BytesIO()
        qrcode.make(url).save(buf, format='PNG')
        qr_img = base64.b64encode(buf.getvalue()).decode()
    return render_template_string(PAGE_QR, qr_img=qr_img, url=url)

@app.route('/register', methods=['GET', 'POST', 'OPTIONS'])
def register():
    """Регистрирует участника по номеру uid и IP-адресу устройства.
    Поддерживает GET (?uid=N) и POST (JSON-тело).
    Блокирует попытку занять номер с другого IP-адреса."""
    if request.method == 'OPTIONS':
        return '', 204  # ответ на preflight-запрос CORS
    if request.method == 'GET':
        uid_str = request.args.get('uid', '-1')
    else:
        data    = request.get_json(force=True, silent=True) or {}
        uid_str = str(data.get('uid', request.args.get('uid', '-1')))
    try:
        uid = int(uid_str)
    except:
        return jsonify({'ok': False, 'error': 'Некорректный номер'})
    ip = request.remote_addr
    if uid < 0 or uid > 999:
        return jsonify({'ok': False, 'error': 'Некорректный номер'})
    with S['lock']:
        for u, info in S['users'].items():
            if u == uid and info['ip'] != ip:
                return jsonify({'ok': False, 'error': f'Номер {uid} уже занят'})
        if uid not in S['users']:
            S['users'][uid] = {'ip': ip, 'online': True, 'taps': {},
                               'joined': ts(), 'last_seen': time.time()}
            print(f'  ✓ Участник № {uid} зарегистрирован ({ip})')
            log(f'№ {uid} подключился ({ip})')
        else:
            S['users'][uid].update({'ip': ip, 'online': True})
            print(f'  ↺ Участник № {uid} переподключился ({ip})')
            log(f'№ {uid} переподключился ({ip})')
        S['ip_to_uid'][ip] = uid
        ek = exp_key()
        return jsonify({'ok': True,
                        'tap_count': len(S['users'][uid]['taps'].get(ek, [])),
                        'exp_version': S['exp_version']})

@app.route('/state')
def state_route():
    """Возвращает текущее состояние эксперимента.
    Вызывается клиентом каждые 800 мс (long polling).
    Обновляет признак online и метку last_seen участника."""
    uid = int(request.args.get('uid', 0))
    with S['lock']:
        if uid in S['users']:
            S['users'][uid]['online']    = True
            S['users'][uid]['last_seen'] = time.time()
        ek  = exp_key()
        cnt = len(S['users'].get(uid, {}).get('taps', {}).get(ek, []))
        return jsonify({
            'active':        S['active'],
            'exp_version':   S['exp_version'],
            'exp_num':       S['exp_num'],
            'total_exp':     S['total_exp'],
            'video_num':     S['video_num'],
            'screen_format': S['screen_format'],
            'tap_count':     cnt,
        })

@app.route('/tap', methods=['POST'])
def tap_route():
    """Принимает оценку от участника.
    Вычисляет таймкод от начала воспроизведения, сопоставляет его
    с метаданными чанка и сохраняет запись в структуру оценок участника."""
    data = request.get_json()
    uid  = int(data.get('uid', 0))
    with S['lock']:
        if not S['active']:
            return jsonify({'ok': False, 'reason': 'inactive'})
        if uid not in S['users']:
            return jsonify({'ok': False, 'reason': 'not_registered'})
        ek   = exp_key()
        taps = S['users'][uid]['taps'].get(ek, [])
        if len(taps) >= 4:
            return jsonify({'ok': False, 'reason': 'max_reached'})
        tc  = time.time() - S['timer_start'] if S['timer_start'] else 0
        mos = MOS_SCORES[len(taps)]
        ch  = find_chunk(S['video_num'] or 1, tc)
        tap = {
            'mos':          mos,
            'timecode_sec': round(tc, 2),
            'timecode':     mmss(tc),
            'resolution':   ch.get('Разрешение', '—'),
            'scale':        ch.get('Масштаб_%', '—'),
            'bitrate':      ch.get('Битрейт_кбит', '—'),
            'fps':          ch.get('FPS', '—'),
        }
        taps.append(tap)
        S['users'][uid]['taps'][ek] = taps
        log(f'№ {uid} → MOS {mos} в {mmss(tc)}')
        return jsonify({'ok': True, 'tap_count': len(taps),
                        'mos': mos, 'timecode': mmss(tc)})

def run_flask():
    """Запускает Flask в фоновом потоке. Уровень журналирования werkzeug понижен."""
    import logging
    logging.getLogger('werkzeug').setLevel(logging.ERROR)
    app.run(host='0.0.0.0', port=PORT, debug=False, use_reloader=False, threaded=True)

def monitor_connections():
    """Фоновый поток мониторинга отключений участников.
    Участник считается отключённым при отсутствии polling-запросов
    в течение TIMEOUT секунд (закрытие браузера, потеря сети)."""
    TIMEOUT = 20.0
    while True:
        time.sleep(2)
        with S['lock']:
            now = time.time()
            for uid, u in S['users'].items():
                was_online = u['online']
                is_online  = (now - u.get('last_seen', now)) < TIMEOUT
                if was_online and not is_online:
                    u['online'] = False
                    log(f'№ {uid} отключился ({u["ip"]})')
                    print(f'  [{ts()}] ⚠ Участник № {uid} отключился ({u["ip"]})')

# ============================================================
# 9. ВОСПРОИЗВЕДЕНИЕ ВИДЕО (FFplay)
# ============================================================
def launch_ffplay(path):
    """Запускает FFplay на втором мониторе (AP7_Titanium).
    Окно позиционируется по X-координате второго монитора,
    затем разворачивается в полноэкранный режим.
    Параметр -autoexit завершает процесс по окончании файла."""
    vf  = "drawtext=text='%{pts\\:hms}':fontsize=40:fontcolor=white:box=1:boxcolor=black@0.6:x=10:y=10"
    cmd = [
        FFPLAY_PATH,
        '-autoexit',
        '-left', str(SECOND_MONITOR_X),  # X-координата второго монитора
        '-top',  str(SECOND_MONITOR_Y),  # Y-координата второго монитора (12px смещение)
        '-fs',                            # полноэкранный режим на этом мониторе
        '-vf', vf,
        path
    ]
    return subprocess.Popen(cmd, stdin=subprocess.DEVNULL,
                            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

def kill_ffplay(proc):
    """Принудительно завершает процесс FFplay с ожиданием его остановки."""
    if proc and proc.poll() is None:
        proc.terminate()
        try:    proc.wait(timeout=2)
        except: proc.kill()

# ============================================================
# 10. КОНСОЛЬНЫЙ ИНТЕРФЕЙС ОРГАНИЗАТОРА
# ============================================================
def print_users():
    """Выводит таблицу зарегистрированных участников с IP и статусом подключения."""
    with S['lock']:
        users = dict(S['users'])
    if not users:
        print('  (нет участников)'); return
    print(f'  {"─"*48}')
    print(f'  {"№":^5} │ {"IP":^15} │ {"Статус":^10} │ {"Вход":^7}')
    print(f'  {"─"*48}')
    for uid in sorted(users):
        u  = users[uid]
        st = '✓ онлайн' if u['online'] else '⚠ офлайн'
        print(f'  {uid:^5} │ {u["ip"]:^15} │ {st:^10} │ {u["joined"]:^7}')
    print(f'  {"─"*48}')

def live_status(proc):
    """Отображает в реальном времени таймкод и прогресс оценок каждого участника.
    Обновляется каждые 500 мс до завершения воспроизведения."""
    ek = exp_key()
    while proc.poll() is None:
        with S['lock']:
            users = dict(S['users'])
            t0    = S['timer_start']
        el   = time.time() - t0 if t0 else 0
        bars = '  '.join(
            f'№{uid} {"●"*len(users[uid]["taps"].get(ek,[]))+"○"*(4-len(users[uid]["taps"].get(ek,[])))}'
            for uid in sorted(users)
        ) or 'ждём участников...'
        print(f'\r  ▶ {mmss(el)}  │  {bars}    ', end='', flush=True)
        time.sleep(0.5)
    print()

def print_results(ek, exp_num, video_num, fmt):
    """Выводит таблицу оценок всех участников по завершении просмотра.
    Строки без зафиксированных оценок отображаются со значением None."""
    with S['lock']:
        users = dict(S['users'])
    sep = '═'*70
    print(f'\n  {sep}')
    print(f'  РЕЗУЛЬТАТЫ — Эксперимент {exp_num} │ Видео {video_num} │ {fmt}')
    print(f'  {"─"*70}')
    print(f'  {"Участник":^10} │ {"MOS":^5} │ {"Таймкод":^8} │ '
          f'{"Разрешение":^12} │ {"Масштаб":^8} │ {"Битрейт":^10}')
    print(f'  {"─"*70}')
    for uid in sorted(users):
        taps = users[uid]['taps'].get(ek, [])
        for i, mos in enumerate(MOS_SCORES):
            lbl = str(uid) if i == 0 else ''
            if i < len(taps):
                t = taps[i]
                print(f'  {lbl:^10} │ {mos:^5} │ {t["timecode"]:^8} │ '
                      f'{str(t["resolution"]):^12} │ {str(t["scale"]):^8} │ '
                      f'{str(t["bitrate"]):^10}')
            else:
                print(f'  {lbl:^10} │ {mos:^5} │ {"None":^8} │ '
                      f'{"—":^12} │ {"—":^8} │ {"—":^10}')
        print(f'  {"·"*70}')
    print(f'  {sep}')

# ============================================================
# 11. ЗАПИСЬ РЕЗУЛЬТАТОВ
# ============================================================
def save_csv(ek, video_num, fmt, sid):
    """Дописывает результаты одного эксперимента в CSV-файл текущей сессии.
    Для участников без зафиксированных оценок записывается значение None."""
    with S['lock']:
        users = dict(S['users'])
    exists = os.path.exists(RESULTS_CSV)
    fields = ['Сессия','Испытуемый','Видео','Формат_экрана',
              'MOS','Таймкод','Разрешение','Масштаб_%',
              'Битрейт_кбит','FPS','Дата_время']
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with open(RESULTS_CSV, 'a', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=fields)
        if not exists:
            w.writeheader()
        for uid in sorted(users):
            taps = users[uid]['taps'].get(ek, [])
            for i, mos in enumerate(MOS_SCORES):
                if i < len(taps):
                    t = taps[i]
                    row = {'Сессия': sid, 'Испытуемый': uid,
                           'Видео': video_num, 'Формат_экрана': fmt,
                           'MOS': mos, 'Таймкод': t['timecode'],
                           'Разрешение': t['resolution'], 'Масштаб_%': t['scale'],
                           'Битрейт_кбит': t['bitrate'], 'FPS': t['fps'],
                           'Дата_время': now}
                else:
                    row = {'Сессия': sid, 'Испытуемый': uid,
                           'Видео': video_num, 'Формат_экрана': fmt,
                           'MOS': mos, 'Таймкод': 'None', 'Разрешение': 'None',
                           'Масштаб_%': 'None', 'Битрейт_кбит': 'None',
                           'FPS': 'None', 'Дата_время': now}
                w.writerow(row)
            f.write('\n')  # пустая строка-разделитель между участниками

def save_xlsx(sid):
    """Создаёт Excel-файл с объединёнными ячейками (Сессия, Участник, Видео, Формат),
    жирными разделителями между блоками форматов и участников,
    цветовым кодированием строк по значению MOS."""
    if not os.path.exists(RESULTS_CSV):
        return
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from itertools import groupby

        df = pd.read_csv(RESULTS_CSV, encoding='utf-8-sig', keep_default_na=False)
        if df.empty:
            return
        xlsx_path = RESULTS_CSV.replace('.csv', '.xlsx')
        wb = Workbook(); ws = wb.active; ws.title = 'Результаты'

        # ── Цветовая схема ──────────────────────────────────────────
        MOS_CLR = {5:'D5E8D4', 4:'DAE8FC', 3:'FFF2CC', 2:'F8CECC'}
        C_HEAD  = '1F3864'  # шапка столбцов (тёмно-синий)
        C_TITLE = '2E4057'  # заголовок листа
        C_USER  = 'CFD8DC'  # блок участника (серо-голубой)
        C_FMT   = 'EDE7F6'  # блок формата экрана (светло-фиолетовый)
        C_NONE  = 'F5F5F5'  # строка без данных
        C_SEP   = '37474F'  # разделитель между участниками

        def fl(h):  return PatternFill('solid', fgColor=h)
        def sd(w='thin', c='AAAAAA'): return Side(style=w, color=c)
        def brd(lw='thin', tw='thin', rw='thin', bw='thin'):
            return Border(left=sd(lw), top=sd(tw), right=sd(rw), bottom=sd(bw))
        def thick():
            s = sd('medium', '333333')
            return Border(left=s, top=s, right=s, bottom=s)

        cen  = Alignment(horizontal='center', vertical='center')
        cen2 = Alignment(horizontal='center', vertical='center', wrap_text=True)
        lft  = Alignment(horizontal='left',   vertical='center')
        COL  = get_column_letter
        NC   = 11  # число столбцов

        # ── Строка 1: заголовок листа ────────────────────────────────
        ws.merge_cells(f'A1:{COL(NC)}1')
        c = ws.cell(1, 1, value='РЕЗУЛЬТАТЫ СУБЪЕКТИВНОГО ЭКСПЕРИМЕНТА')
        c.font=Font(bold=True, size=14, color='FFFFFF')
        c.fill=fl(C_TITLE); c.alignment=cen
        c.border=brd('medium','medium','medium','medium')
        ws.row_dimensions[1].height = 26

        # ── Строка 2: шапка столбцов ─────────────────────────────────
        hdrs   = ['Сессия','Испытуемый','Видео','Формат экрана','MOS',
                  'Таймкод','Разрешение','Масштаб %','Битрейт кбит/с','FPS','Дата и время']
        widths = [20, 12, 7, 12, 6, 9, 13, 10, 14, 7, 20]
        for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
            c = ws.cell(2, ci, value=h)
            c.font=Font(bold=True, size=10, color='FFFFFF')
            c.fill=fl(C_HEAD); c.alignment=cen2
            c.border=brd('medium','medium','medium','medium')
            ws.column_dimensions[COL(ci)].width = w
        ws.row_dimensions[2].height = 30
        ws.freeze_panes = 'A3'

        # ── Подготовка данных ─────────────────────────────────────────
        cols = ['Сессия','Испытуемый','Видео','Формат_экрана','MOS',
                'Таймкод','Разрешение','Масштаб_%','Битрейт_кбит','FPS','Дата_время']
        rows = [list(r) for r in df[cols].itertuples(index=False)]

        def grp(seq, key):
            """Группирует последовательность по ключу, возвращает (ключ, список)."""
            return [(k, list(g)) for k, g in groupby(seq, key=key)]

        ri       = 3    # текущая строка Excel (начиная с 3-й)
        prev_uid = None

        for (sid_v, uid), u_rows in grp(rows, lambda r: (r[0], r[1])):
            # ── Разделитель между участниками ────────────────────────
            if prev_uid is not None:
                ws.row_dimensions[ri].height = 7
                for ci in range(1, NC+1):
                    c = ws.cell(ri, ci); c.fill=fl(C_SEP)
                    c.border=brd('medium','medium','medium','medium')
                ri += 1

            u_start = ri
            u_total = len(u_rows)
            prev_fmt = None

            for (fmt,), f_rows in grp(u_rows, lambda r: (r[3],)):
                f_start = ri
                f_total = len(f_rows)
                is_new_fmt = (prev_fmt is not None)  # признак смены формата

                for (vid,), v_rows in grp(f_rows, lambda r: (r[2],)):
                    v_start = ri
                    for row in v_rows:
                        mos     = row[4]
                        is_none = str(row[5]) == 'None'
                        rc      = C_NONE if is_none else MOS_CLR.get(
                                    int(mos) if str(mos).isdigit() else 0, 'FFFFFF')
                        # жирная верхняя граница при смене формата экрана
                        top_w = 'medium' if (ri == f_start and is_new_fmt) else 'thin'
                        top_c = '444444' if top_w == 'medium' else 'AAAAAA'

                        for ci, val in enumerate(row, 1):
                            # ячейки, которые будут объединены — оставляем пустыми
                            skip = ((ci == 1 and ri > u_start) or   # Сессия
                                    (ci == 2 and ri > u_start) or   # Участник
                                    (ci == 4 and ri > f_start) or   # Формат
                                    (ci == 3 and ri > v_start))     # Видео
                            c = ws.cell(ri, ci, value=(None if skip else val))
                            c.fill=fl(rc); c.font=Font(size=10)
                            c.alignment = lft if ci == NC else cen
                            c.border = Border(
                                left=sd(), right=sd(), bottom=sd(),
                                top=sd(top_w, top_c))
                        ri += 1

                    # ── Объединяем ячейки столбца «Видео» ────────────
                    if len(v_rows) > 1:
                        ws.merge_cells(f'{COL(3)}{v_start}:{COL(3)}{ri-1}')
                        mc = ws.cell(v_start, 3)
                        mc.alignment = cen
                        mc.font = Font(size=10, bold=True)

                # ── Объединяем ячейки столбца «Формат экрана» ────────
                if f_total > 1:
                    ws.merge_cells(f'{COL(4)}{f_start}:{COL(4)}{ri-1}')
                    mc = ws.cell(f_start, 4)
                    mc.alignment = cen
                    mc.font = Font(size=10, bold=True)
                    mc.fill = fl(C_FMT)

                prev_fmt = fmt

            # ── Объединяем ячейки «Сессия» и «Участник» ─────────────
            if u_total > 1:
                for col_i in [1, 2]:
                    ws.merge_cells(f'{COL(col_i)}{u_start}:{COL(col_i)}{ri-1}')
                    mc = ws.cell(u_start, col_i)
                    mc.alignment = cen
                    mc.font = Font(size=10, bold=True)
                    mc.fill = fl(C_USER)
                    mc.border = thick()

            prev_uid = uid

        wb.save(xlsx_path)
        print(f'  ✓ Excel сохранён: {xlsx_path}')
    except Exception as e:
        import traceback
        print(f'  ⚠ Excel не создан: {e}')
        traceback.print_exc()

def delete_csv(video_num, fmt, sid):
    """Удаляет строки повторяемого эксперимента из CSV для последующей перезаписи."""
    if not os.path.exists(RESULTS_CSV):
        return
    try:
        df   = pd.read_csv(RESULTS_CSV, encoding='utf-8-sig')
        mask = ~((df['Сессия']==sid)&(df['Видео']==video_num)&(df['Формат_экрана']==fmt))
        df[mask].to_csv(RESULTS_CSV, index=False, encoding='utf-8-sig')
    except Exception as e:
        print(f'  ⚠ Ошибка удаления: {e}')

def finalize_none(ek):
    """Инициализирует пустой список оценок для участников, не нажавших кнопку,
    чтобы гарантировать запись строк None в CSV для каждого участника."""
    with S['lock']:
        for u in S['users'].values():
            if ek not in u['taps']:
                u['taps'][ek] = []

# ============================================================
# 12. ГЛАВНАЯ ФУНКЦИЯ
# ============================================================
def main():
    """Реализует полный протокол сессии эксперимента:
    регистрация участников → настройка параметров →
    итерация по форматам экрана и видео → сохранение результатов."""
    global RESULTS_CSV  # путь задаётся динамически после создания папки сессии

    load_chunks()
    sid = datetime.now().strftime('%Y%m%d %H%M%S')  # идентификатор сессии
    url = f'http://{get_local_ip()}:{PORT}'

    # Проверка подключения второго монитора (AP7_Titanium / DISPLAY2)
    print('  Проверка второго монитора...')
    if check_second_monitor():
        print(f'  ✓ Монитор DISPLAY2 подключён (X={SECOND_MONITOR_X}, Y={SECOND_MONITOR_Y})')
        print(f'  ✓ Видео будет выводиться на второй экран')
    else:
        print(f'  ⚠ Монитор DISPLAY2 не обнаружен!')
        print(f'  → Подключите монитор через HDMI и нажмите Enter для повторной проверки,')
        print(f'    или e+Enter чтобы продолжить без второго монитора (видео на основном экране)')
        while True:
            cmd = input('  → ').strip().lower()
            if cmd == 'e':
                print('  ⚠ Продолжаем без второго монитора')
                break
            if check_second_monitor():
                print(f'  ✓ Монитор DISPLAY2 подключён')
                break
            print('  ⚠ Всё ещё не обнаружен. Enter — проверить снова, e — продолжить')

    # Создание нумерованной папки для файлов текущей сессии
    existing  = [d for d in os.listdir(RESULTS_DIR)
                 if d.startswith('Сессия ') and
                 os.path.isdir(os.path.join(RESULTS_DIR, d))]
    sess_nums = []
    for d in existing:
        try: sess_nums.append(int(d.split(' ')[1]))
        except: pass
    sess_num = (max(sess_nums) + 1) if sess_nums else 1
    sess_dir = os.path.join(RESULTS_DIR, f'Сессия {sess_num}')
    os.makedirs(sess_dir, exist_ok=True)
    RESULTS_CSV = os.path.join(sess_dir, 'Результаты экспериментов.csv')
    print(f'  ✓ Папка сессии: Сессия {sess_num}  ({sess_dir})')

    prev = max_uid_in_csv()
    print(f'''
╔══════════════════════════════════════════════════════╗
║       СИСТЕМА СУБЪЕКТИВНОЙ ОЦЕНКИ ВИДЕО (Web)       ║
╠══════════════════════════════════════════════════════╣
║  Участники:   {url:<38}║
║  QR страница: {(url+"/qr"):<38}║
║  → Откройте QR страницу на большом экране           ║
╠══════════════════════════════════════════════════════╣
║  УПРАВЛЕНИЕ:                                        ║
║  Enter    — продолжить / подтвердить                ║
║  r+Enter  — повторить эксперимент                   ║
║  q+Enter  — пропустить                              ║
║  e+Enter  — выйти из программы                      ║
╚══════════════════════════════════════════════════════╝''')

    if prev:
        print(f'\n  ℹ Предыдущих участников: {prev}  (новые номера от {prev+1})')

    # Ожидание подключений с живым выводом событий регистрации
    def watch_registrations():
        seen = set()
        while not watch_registrations.stop:
            with S['lock']:
                events = list(S['events']); S['events'].clear()
                users  = dict(S['users'])
            for e in events: print(e)
            current = set(users.keys())
            if current != seen:
                seen = current
                if seen:
                    print(f'\r  Подключено: {len(seen)} участников — {sorted(seen)}   ', flush=True)
            time.sleep(0.5)
    watch_registrations.stop = False
    wt = threading.Thread(target=watch_registrations, daemon=True)
    wt.start()
    print('\n  Ожидаем участников... Когда все готовы → Enter')
    input('  → ')
    watch_registrations.stop = True; wt.join(timeout=1)
    print(); print_users()
    session_start = time.time()  # метка начала сессии для подсчёта длительности

    # Настройка параметров: количество и порядок видео
    print(f'\n  {"═"*50}\n  НАСТРОЙКА СЕССИИ\n  {"═"*50}')
    print('\n  Сколько видео? (2/3/4):')
    n     = int(ask('  → ', ['2','3','4']))
    print('\n  Порядок: (1) последовательно  (2) случайный:')
    order = ask('  → ', ['1','2'])
    vord  = list(range(1,5))[:n] if order=='1' else random.sample(list(range(1,5)), n)
    total = n * 3  # число экспериментов = число видео × число форматов экрана
    with S['lock']:
        S['total_exp'] = total
    print(f'\n  Видео: {vord}  │  Форматы: 4K → 1080 → 720  │  Экспериментов: {total}')
    print('\n  Enter для старта...'); input('  → ')

    exp_num = 0; done = False

    for fi, fmt in enumerate(FORMATS):
        if done: break

        # Ввод формата и автоматическая смена разрешения дисплея
        print(f'\n  {"═"*50}\n  ФОРМАТ ЭКРАНА {fi+1}/3\n  {"═"*50}')
        print(f'  Введите формат ({"/".join(FORMATS)}):')
        fi_raw = ask('  → ', ['4k','4к','1080','720'])
        fmt    = '4K' if fi_raw in ('4k','4к') else fi_raw.upper()
        ok, msg = change_resolution(fmt)
        print(f'  {"✓" if ok else "⚠"} {msg}')
        if not ok:
            print('  Смените вручную, затем Enter'); input('  → ')
        print(f'\n  Убедитесь что экран {fmt}. Enter — начинаем  │  e — выйти')
        if input('  → ').strip().lower() == 'e': break
        with S['lock']:
            S['screen_format'] = fmt

        for vn in vord:
            if done: break
            exp_num += 1

            while True:  # цикл повтора одного эксперимента (по команде r)
                # Сброс состояния перед каждым запуском эксперимента
                with S['lock']:
                    S['exp_version'] += 1; S['exp_num'] = exp_num
                    S['video_num'] = vn; S['active'] = False; S['timer_start'] = None
                    ek = exp_key()
                    for u in S['users'].values():
                        u['taps'].pop(ek, None)  # очищаем тыки текущей версии

                ek    = exp_key()
                vpath = os.path.join(VIDEO_DIR, VIDEO_MAP[vn])
                print(f'\n  {"═"*50}')
                print(f'  ЭКСПЕРИМЕНТ {exp_num}/{total} │ Видео {vn} │ {fmt}')
                print(f'  {"─"*50}')
                print_users()
                print('\n  Enter — запустить  │  q — пропустить  │  e — выйти')
                cmd = input('  → ').strip().lower()
                if cmd == 'e': done = True; break
                if cmd == 'q':
                    finalize_none(ek); save_csv(ek, vn, fmt, sid)
                    print('  ⏭ Пропущено'); break

                # Активация кнопок с буфером для синхронизации polling-цикла
                print('\n  Активирую кнопки...')
                with S['lock']: S['active'] = True
                time.sleep(2.5)  # ожидание первого polling-ответа от всех устройств
                print('  ✓ Кнопки активны. Запуск видео...')
                time.sleep(0.5)

                proc = launch_ffplay(vpath)
                with S['lock']: S['timer_start'] = time.time()
                live_status(proc)  # блокирующий вывод статуса до конца воспроизведения

                with S['lock']: S['active'] = False
                kill_ffplay(proc); finalize_none(ek)
                print_results(ek, exp_num, vn, fmt)

                print(f'\n  Прогресс: {exp_num}/{total}')
                print('  Enter — сохранить  │  r — повторить  │  q — пропустить  │  e — выйти')
                cmd = input('  → ').strip().lower()
                if cmd == 'e':
                    done = True; save_csv(ek, vn, fmt, sid); break
                elif cmd == 'r':
                    delete_csv(vn, fmt, sid); print('  ↺ Повтор...'); continue
                elif cmd == 'q':
                    print('  ⏭ Пропущено'); break
                else:
                    save_csv(ek, vn, fmt, sid); save_xlsx(sid)
                    print(f'  ✓ Сохранено → {RESULTS_CSV}'); break

        if not done and fi < 2:
            print(f'\n  ✓ Формат {fmt} завершён ({exp_num}/{total})')
            print('  Смените экран. Enter — продолжить  │  e — выйти')
            if input('  → ').strip().lower() == 'e': break

    with S['lock']: S['active'] = False

    # Вывод итогов сессии
    elapsed    = time.time() - session_start
    mins, secs = divmod(int(elapsed), 60)
    print(f'\n  {"═"*50}')
    print(f'  СЕССИЯ ЗАВЕРШЕНА  ({sid})')
    print(f'  Длительность: {mins:02d}:{secs:02d}')
    print(f'  Папка: Сессия {sess_num}')
    print(f'  Результаты: {RESULTS_CSV}')
    print(f'  {"═"*50}\n')

    # Возврат разрешения дисплея по выбору организатора
    print('  Вернуть разрешение экрана? Введите формат (4K/1080) или Enter чтобы пропустить:')
    res_choice = input('  → ').strip().lower()
    if res_choice in ('4k', '4к', '1080'):
        ok, msg = change_resolution('4K' if res_choice in ('4k','4к') else '1080')
        print(f'  {"✓" if ok else "⚠"} {msg}')

# ============================================================
# 13. ТОЧКА ВХОДА
# ============================================================
if __name__ == '__main__':
    try:
        import flask
    except ImportError:
        print('Установи: pip install flask qrcode[pil]'); sys.exit(1)
    if not os.path.exists(FFPLAY_PATH):
        print(f'FFplay не найден: {FFPLAY_PATH}'); sys.exit(1)

    # Автоматическое открытие порта в брандмауэре Windows
    try:
        rule_name = 'Flask_VideoExperiment_5000'
        check = subprocess.run(
            ['netsh','advfirewall','firewall','show','rule','name='+rule_name],
            capture_output=True, text=True)
        if 'No rules match' in check.stdout or not check.stdout.strip():
            r = subprocess.run([
                'netsh','advfirewall','firewall','add','rule',
                f'name={rule_name}', 'dir=in', 'action=allow',
                'protocol=TCP', f'localport={PORT}'
            ], capture_output=True, text=True)
            print(f'  {"✓" if r.returncode==0 else "⚠"} Порт {PORT} '
                  f'{"открыт" if r.returncode==0 else "не открыт"} в Windows Firewall')
        else:
            print(f'  ✓ Правило Firewall уже есть (порт {PORT} открыт)')
    except Exception as e:
        print(f'  ⚠ Firewall: {e}')

    # Запуск Flask и монитора подключений в фоновых потоках
    threading.Thread(target=run_flask, daemon=True).start()
    threading.Thread(target=monitor_connections, daemon=True).start()
    time.sleep(1)  # ожидание инициализации Flask перед запуском консоли
    main()
